#!/usr/bin/env python3
"""
evidence_gen.py - AI投委会证据表生成器
读取 data_bundle.json，生成 8-Sheet Excel 证据表。
用法: python3 evidence_gen.py <data_bundle.json> <output.xlsx>
依赖: pip install openpyxl
"""
import sys, json, os
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.stderr.write("[ERROR] openpyxl not installed. Run: pip install openpyxl\n")
    sys.exit(1)

# ── 样式常量 ──────────────────────────────────────────────
HEADER_FONT = Font(name="Arial", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGN = Alignment(vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
ALT_FILL = PatternFill(start_color="f2f7fb", end_color="f2f7fb", fill_type="solid")


def style_sheet(ws, headers, rows):
    """Apply headers, rows, and formatting to a worksheet."""
    # Headers
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    # Data rows
    for r_idx, row in enumerate(rows, 2):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = CELL_ALIGN
            cell.border = THIN_BORDER
            if r_idx % 2 == 0:
                cell.fill = ALT_FILL
    # Auto-width (capped at 50)
    for col_idx in range(1, len(headers) + 1):
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)
    ws.freeze_panes = "A2"


def ts():
    """Current timestamp string."""
    return datetime.now().strftime("%Y-%m-%d %H:%M")


# ── Sheet 构建器 ──────────────────────────────────────────

def build_financial(wb, data):
    """Sheet 1: 财务数据"""
    ws = wb.active
    ws.title = "财务数据"
    headers = ["指标", "最新值", "上一年", "上两年", "数据来源", "采集时间"]
    fin = data.get("financial", {})
    rows = []
    for key, label in [
        ("revenue", "营业收入(亿元)"), ("net_profit", "归母净利润(亿元)"),
        ("gross_margin", "毛利率(%)"), ("net_margin", "净利率(%)"),
        ("roe", "ROE(%)"), ("roa", "ROA(%)"),
        ("debt_ratio", "资产负债率(%)"), ("current_ratio", "流动比率"),
        ("ocf", "经营现金流(亿元)"), ("fcf", "自由现金流(亿元)"),
    ]:
        v = fin.get(key, {})
        if isinstance(v, dict):
            rows.append([label, v.get("latest", "-"), v.get("y1", "-"),
                         v.get("y2", "-"), v.get("source", "-"), ts()])
        else:
            rows.append([label, v if v else "-", "-", "-", fin.get("source", "-"), ts()])
    if not rows:
        rows = [["(无数据)", "-", "-", "-", "-", ts()]]
    style_sheet(ws, headers, rows)


def build_valuation(wb, data):
    """Sheet 2: 估值对比"""
    ws = wb.create_sheet("估值对比")
    headers = ["指标", "当前值", "5年均值", "行业中位数", "数据来源", "采集时间"]
    val = data.get("valuation", {})
    rows = []
    for key, label in [
        ("pe_ttm", "PE(TTM)"), ("pb", "PB"), ("ps", "PS"),
        ("peg", "PEG"), ("ev_ebitda", "EV/EBITDA"),
        ("div_yield", "股息率(%)"), ("market_cap", "总市值(亿元)"),
    ]:
        v = val.get(key, {})
        if isinstance(v, dict):
            rows.append([label, v.get("current", "-"), v.get("avg_5y", "-"),
                         v.get("industry_median", "-"), v.get("source", "-"), ts()])
        else:
            rows.append([label, v if v else "-", "-", "-", val.get("source", "-"), ts()])
    if not rows:
        rows = [["(无数据)", "-", "-", "-", "-", ts()]]
    style_sheet(ws, headers, rows)


def build_product_customer(wb, data):
    """Sheet 3: 产品客户"""
    ws = wb.create_sheet("产品客户")
    headers = ["类别", "名称/内容", "占比(%)", "趋势", "数据来源", "采集时间"]
    rows = []
    fin = data.get("financial", {})
    products = fin.get("products", [])
    for p in products:
        rows.append(["产品", p.get("name", "-"), p.get("revenue_pct", "-"),
                     p.get("trend", "-"), p.get("source", "-"), ts()])
    customers = fin.get("top_customers", [])
    for c in customers:
        rows.append(["客户", c.get("name", "-"), c.get("revenue_pct", "-"),
                     c.get("trend", "-"), c.get("source", "-"), ts()])
    if not rows:
        rows = [["(无数据)", "-", "-", "-", "-", ts()]]
    style_sheet(ws, headers, rows)


def build_shareholders(wb, data):
    """Sheet 4: 股东结构"""
    ws = wb.create_sheet("股东结构")
    headers = ["股东名称", "持股比例(%)", "持股数(万股)", "变动", "身份", "数据来源", "采集时间"]
    sh = data.get("shareholders", {})
    rows = []
    for s in sh.get("top10", []):
        rows.append([s.get("name", "-"), s.get("pct", "-"), s.get("shares", "-"),
                     s.get("change", "-"), s.get("identity", "-"),
                     s.get("source", "-"), ts()])
    inst = sh.get("institutional", {})
    if inst:
        rows.append(["机构合计", inst.get("total_pct", "-"), inst.get("total_shares", "-"),
                     inst.get("change", "-"), "机构", inst.get("source", "-"), ts()])
    if not rows:
        rows = [["(无数据)", "-", "-", "-", "-", "-", ts()]]
    style_sheet(ws, headers, rows)


def build_risks(wb, data):
    """Sheet 5: 风险清单"""
    ws = wb.create_sheet("风险清单")
    headers = ["风险类别", "风险描述", "影响程度", "发生概率", "数据来源", "采集时间"]
    risk = data.get("risk_events", [])
    rows = []
    for r in risk:
        rows.append([r.get("category", "-"), r.get("description", "-"),
                     r.get("impact", "-"), r.get("probability", "-"),
                     r.get("source", "-"), ts()])
    if not rows:
        rows = [["(无数据)", "-", "-", "-", "-", ts()]]
    style_sheet(ws, headers, rows)


def build_technical(wb, data):
    """Sheet 6: 技术面"""
    ws = wb.create_sheet("技术面")
    headers = ["指标", "值", "信号", "数据来源", "采集时间"]
    tech = data.get("technical", {})
    rows = []
    for key, label in [
        ("price", "当前股价"), ("ma20", "MA20"), ("ma60", "MA60"),
        ("ma120", "MA120"), ("ma250", "MA250"),
        ("rsi", "RSI(14)"), ("macd", "MACD"), ("kdj", "KDJ"),
        ("volume_ratio", "量比"), ("turnover_rate", "换手率(%)"),
        ("boll_upper", "布林上轨"), ("boll_lower", "布林下轨"),
    ]:
        v = tech.get(key, {})
        if isinstance(v, dict):
            rows.append([label, v.get("value", "-"), v.get("signal", "-"),
                         v.get("source", "-"), ts()])
        else:
            rows.append([label, v if v else "-", "-", tech.get("source", "-"), ts()])
    if not rows:
        rows = [["(无数据)", "-", "-", "-", ts()]]
    style_sheet(ws, headers, rows)


def build_sentiment(wb, data):
    """Sheet 7: 舆情"""
    ws = wb.create_sheet("舆情")
    headers = ["来源", "标题/内容", "情感倾向", "热度", "发布时间", "数据来源", "采集时间"]
    ns = data.get("news_sentiment", {})
    rows = []
    for item in ns.get("items", []):
        rows.append([item.get("platform", "-"), item.get("title", "-"),
                     item.get("sentiment", "-"), item.get("heat", "-"),
                     item.get("published_at", "-"), item.get("source", "-"), ts()])
    if not rows:
        rows = [["(无数据)", "-", "-", "-", "-", "-", ts()]]
    style_sheet(ws, headers, rows)


def build_scores(wb, data):
    """Sheet 8: 评分汇总"""
    ws = wb.create_sheet("评分汇总")
    headers = ["维度", "权重(%)", "得分(0-100)", "加权分", "关键依据", "采集时间"]
    # 从 analyses 中读取（如果 data_bundle 中嵌入了 analyses）
    scores = data.get("scores", {})
    weights = {
        "industry_logic": 20, "company_quality": 20, "growth": 20,
        "financial_health": 15, "valuation": 15, "risk_reward": 10,
    }
    labels = {
        "industry_logic": "产业逻辑", "company_quality": "公司质地",
        "growth": "成长性", "financial_health": "财务健康",
        "valuation": "估值合理性", "risk_reward": "风险收益比",
    }
    rows = []
    total = 0
    for k, w in weights.items():
        s = scores.get(k, {})
        score = s.get("score", "-")
        rationale = s.get("rationale", "-")
        weighted = round(score * w / 100, 2) if isinstance(score, (int, float)) else "-"
        if isinstance(weighted, (int, float)):
            total += weighted
        rows.append([labels[k], w, score, weighted, rationale, ts()])
    rows.append(["合计", 100, "-", round(total, 2), "-", ts()])
    style_sheet(ws, headers, rows)


# ── 主函数 ────────────────────────────────────────────────

def main():
    if len(sys.argv) != 3:
        sys.stderr.write("Usage: python3 evidence_gen.py <data_bundle.json> <output.xlsx>\n")
        sys.exit(1)

    input_path = sys.argv[1]
    output_path = sys.argv[2]

    if not os.path.exists(input_path):
        sys.stderr.write(f"[ERROR] File not found: {input_path}\n")
        sys.exit(1)

    with open(input_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    wb = Workbook()
    build_financial(wb, data)
    build_valuation(wb, data)
    build_product_customer(wb, data)
    build_shareholders(wb, data)
    build_risks(wb, data)
    build_technical(wb, data)
    build_sentiment(wb, data)
    build_scores(wb, data)

    wb.save(output_path)
    print(f"[OK] Evidence Excel saved: {output_path}")
    print(f"     Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
